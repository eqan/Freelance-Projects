import openpyxl
from os.path import exists
import pandas as pd


def generateMasterList():
    wb_obj = openpyxl.load_workbook("./PrefixSuffix.xlsx")
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    headers = ['Domains']
    workbook_name = 'MasterList.xlsx'
    wb = openpyxl.Workbook()
    page = wb.active
    page.title = 'Master List'
    page.append(headers)
    
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        cell_obj2 = sheet_obj.cell(row = i, column = 2)
        if cell_obj.value is None:
            break
        page.append([cell_obj.value + cell_obj2.value])
        page.append([cell_obj2.value + cell_obj.value])

    wb.save(filename = workbook_name)

if not exists('./MasterList.xlsx'):
    generateMasterList()
else:
    print("Master List Already exists, delete the Master List and History log file to recompute the list!")

# rows = [
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# ]


df_data = pd.read_excel('./History.xlsx')
df_data_2 = pd.read_excel('./MasterList.xlsx')
for index, row in df_data_2.iterrows():
    new_row = {'Domain':row[0]}
    df_data = df_data.append(new_row, ignore_index=True)


print(df_data)

# writing to Excel
df_result = pd.ExcelWriter('./History.xlsx')
# write students data to excel
df_data.to_excel(df_result,index = False)
# save the students result excel
df_result.save()


# book = openpyxl.Workbook()
# sheet = book.active
# sheet_max_row = sheet.max_row

# rows = [
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# ]

# # sheet_obj = wb_obj.active
# for row in rows:
#     sheet.append(row)
# m_row = sheet_obj.max_row
# for i in range(m_row+1, m_row + 1 + len(rows)):
#     cell_obj = sheet_obj.cell(row = i, column = 1)
#     if cell_obj.value is None:
#         sheet.append(rows[len(rows)-1])
#         rows.pop()


# wb_obj = openpyxl.load_workbook("./MasterList.xlsx")
# sheet_obj = wb_obj.active
# m_row = sheet_obj.max_row
# for i in range(2, m_row + 1):
#     cell_obj = sheet_obj.cell(row = i, column = 1)
#     if cell_obj.value is None:
#         for j in range(2, sheet_max_row + 1):
#             cell_obj_2 = sheet_obj.cell(row = i, column = 1)
#             if cell_obj_2 is None:
#                 for row in rows:
#                     sheet.append(row)
#         break

# book.save('History.xlsx')