import time
import os
import openpyxl
import pandas as pd
from datetime import date
from os.path import exists
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

chromeDriverPath = './chromedriver'
url = "https://www.namebase.io/domains?page=1#marketplace"

def generateMasterList():
    prefixSuffixObj = openpyxl.load_workbook("./PrefixSuffix.xlsx")
    prefixSuffixSheetObj = prefixSuffixObj.active
    prefixSuffixMaxRow = prefixSuffixSheetObj.max_row

    masterListHeaders = ['Domains', 'Price']
    masterListBookName = 'MasterList.xlsx'
    wb = openpyxl.Workbook()
    page = wb.active
    page.title = 'Master List'
    page.append(masterListHeaders)
    
    for i in range(2, prefixSuffixMaxRow + 1):
        cell_obj = prefixSuffixSheetObj.cell(row = i, column = 1)
        cell_obj2 = prefixSuffixSheetObj.cell(row = i, column = 2)
        cell_obj3 = prefixSuffixSheetObj.cell(row = i, column = 3)
        if cell_obj.value is None:
            break
        page.append([cell_obj.value + cell_obj2.value, cell_obj3.value])
        page.append([cell_obj2.value + cell_obj.value, cell_obj3.value])

    wb.save(filename = masterListBookName)


print("Importing your data!..")
if not exists('./MasterList.xlsx'):
    generateMasterList()
else:
    print("Master List Already exists, delete the Master List and History log file to recompute the list!")

f = open("credentials.txt", "r")
email = (f.readline()).strip()
password = (f.readline()).strip()
f.close()
print(email)
print(password)

historyData = pd.read_excel('./History.xlsx')
masterListData = pd.read_excel('./MasterList.xlsx')


s = Service('./chromedriver')
driver = webdriver.Chrome(service=s)
driver.maximize_window()

driver.execute_script("window.open()")
driver.switch_to.window(driver.window_handles[1])
driver.get(url)

try:
    time.sleep(2)
    not_logged_in = driver.find_element(by=By.XPATH, value='//*[@href="/login"]')
    not_logged_in.click()
    time.sleep(1)
    try:
        login_area = driver.find_element(by=By.XPATH, value='//*[@placeholder="Email"]').send_keys(email)
        password_area = driver.find_element(by=By.XPATH, value='//*[@placeholder="Password"]').send_keys(password)
        time.sleep(2)
        submit_btn = driver.find_element(by=By.XPATH, value='//*[@type="submit"]').click()
        print('Login Successful')

    except Exception as error:
        print(error)
        print('Login Unsuccessful')

except Exception as error:
    print(error)
    print('Already Logged In')

time.sleep(2)
z = 0
y = 1
for index,(domainName, price) in masterListData.iterrows():
    masterListWB = openpyxl.load_workbook("./MasterList.xlsx")
    masterListSheet = masterListWB.active
    if(domainName != None):
        new_row = {
        'Domain':domainName, 'Status': "", 'Highest Lockout': "", 'Winner': "", 'Originating Sale Price': "",
         'Final Selling Price': "", 'Namebase Commission': "", "Total Payout": "", "Inserted Date": ""
        }
        print(f'Processing for domain {domainName}')

        url = f'https://www.namebase.io/domains/{domainName}'
        driver.get(url)
        try:
            sellingPrice = driver.find_elements(by=By.XPATH, value="//*[@class='Text__TextStyledElement-sc-9cd9ed-0 fejZys'])")[2]
            new_row["Final Selling Price"] = sellingPrice.text
            # print("sellingPrice: ", sellingPrice.text)
        except Exception as error:
            print(error)
            new_row["Final Selling Price"] = "Selling Price Not Found"
            # print("Selling Price Not Found")

        url = f'https://www.namebase.io/domain-manager/{domainName}'
        driver.get(url)
        time.sleep(2)
        try:
            # currentStatus = driver.find_element(by=By.XPATH, value="//*[@class='Text__TextStyledElement-sc-9cd9ed-0 klnwhI  domainManagerUiComponents__Status-sc-1jci4ma-10 gNIFSj']")
            currentStatus = driver.find_element(by=By.XPATH, value="//*[contains(text(),'up-to-date as of Block')]")
            new_row["Status"] = currentStatus.text
            print("currentStatus: ", str(currentStatus.text))
        except Exception as error:
            print(error)
            new_row["Status"] = "No status found!"
            # print("No status found!")

        try:
            click_sale_button = driver.find_element(by=By.XPATH, value="//*[@class='Switch__Track-xjjcnb-0 deCrCu']").click()
        except Exception as error:
            print(error)
            print("sale button not clicked!")


        try:
            textarea = driver.find_element(by=By.XPATH,value=f"//*[@class='TextInputStyledComponents__OutlinedInput-icv573-3 TextInput___StyledOutlinedInput-pale85-0 fRBNXJ']").send_keys(price)
            textarea = driver.find_element(by=By.XPATH,value=f"//*[@class='TextInputStyledComponents__OutlinedInput-icv573-3 TextInput___StyledOutlinedInput-pale85-0 fRBNXJ']").send_keys(Keys.ENTER)
        except Exception as error:
            print(error)
            print("No text input found!")

        time.sleep(2)

        try:
            valuesTable = driver.find_elements(by=By.XPATH, value="//*[@class='Text__TextStyledElement-sc-9cd9ed-0 hvyzce']")
        except Exception as error:
            print(error)
            print("Values table not found!")

        try:
            highestLockOut = driver.find_elements(by=By.XPATH, value="//*[@class='Text__TextStyledElement-sc-9cd9ed-0 hscVKy']")[1]
            new_row["Highest Lockout"] = highestLockOut.text
            # print("highestLockOut: ", highestLockOut.text)
        except Exception as error:
            print(error)
            new_row["Highest Lockout"] = "Highest Lockout price not found!"
            # print("Highest Lockout price not found!")

        try: 
            originatingSalePrice = valuesTable[1].text
            new_row["Originating Sale Price"] = originatingSalePrice
            # print("originatingSalePrice: ", originatingSalePrice)
        except Exception as error:
            print(error)
            new_row["Originating Sale Price"] = "Couldnt extract originatingSalePrice"
            # print("Couldnt extract originatingSalePrice")

        try: 
            nameBaseCommission = valuesTable[3].text
            new_row["Namebase Commission"] = nameBaseCommission
            # print("nameBaseCommission: ", nameBaseCommission)
        except Exception as error:
            print(error)
            new_row["Namebase Commission"] = "Couldnt extract nameBaseCommission"
            # print("Couldnt extract nameBaseCommission")
        try: 
            payOut = valuesTable[5].text
            new_row["Total Payout"] = payOut
            # print("payOut: ", payOut)
        except Exception as error:
            print(error)
            new_row["Total Payout"] = "Couldnt extract payOut"
            # print("Couldnt extract payOut")
            
        try: 
            click_for_sale = driver.find_element(by=By.XPATH, value="//*[contains(text(),'List for sale')]").click()
            # print("payOut: ", payOut)
        except Exception as error:
            print(error)
            print("Couldnt find sale button!")

        time.sleep(2)

        masterListSheet.delete_rows(masterListSheet.min_row + 1, 1)
        masterListWB.save("./MasterList.xlsx")

        new_row["Winner"] = "Me"
        new_row["Inserted Date"] = date.today()
        historyData = historyData.append(new_row, ignore_index=True)
        df_result = pd.ExcelWriter('./History.xlsx')
        historyData.to_excel(df_result,index = False)
        df_result.save()
        print(new_row)

        z+=1
        driver.execute_script("window.open()")
        driver.switch_to.window(driver.window_handles[z+1])
        driver.get(url)
        driver.switch_to.window(driver.window_handles[z-1])
        driver.close()
        driver.switch_to.window(driver.window_handles[z])
        z-=1
        try:
            print (driver.window_handles[z+1], driver.title, driver.current_url)
        except:
            print("Couldnt print line")
        time.sleep(1)

        try:
            time.sleep(2)
            not_logged_in = driver.find_element(by=By.XPATH, value='//*[@href="/login"]')
            not_logged_in.click()
            time.sleep(1)

            try:
                login_area = driver.find_element(by=By.XPATH, value='//*[@placeholder="Email"]').send_keys(email)
                password_area = driver.find_element(by=By.XPATH, value='//*[@placeholder="Password"]').send_keys(password)
                time.sleep(1)
                submit_btn = driver.find_element(by=By.XPATH, value='//*[@type="submit"]').click()
                print('Login Successful')

            except Exception as error:
                print(error)
                print('Login Unsuccessful')

        except Exception as error:
            print(error)
            print('Already Logged In')

print("All domains Processed!")